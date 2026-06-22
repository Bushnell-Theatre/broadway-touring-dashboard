"""
Broadway Touring Report Processor
Bushnell Center for the Performing Arts

Processes weekly Broadway League XLSX touring reports and outputs a single
clean data.json file for use by the Sales Intelligence Dashboard.

Usage:
    python process_touring.py <input_dir> <output_file>

    input_dir   : folder containing one or more .xlsx touring report files
    output_file : path for the output data.json (e.g. ./data.json)

Example:
    python process_touring.py ./reports ./data.json
"""

import json
import os
import re
import sys
from datetime import datetime, timezone

import openpyxl

# ── CONSTANTS ─────────────────────────────────────────────────────────────────

BUSHNELL_AVG   = 2722
LOWER_BOUND    = BUSHNELL_AVG * 0.9
UPPER_BOUND    = BUSHNELL_AVG * 1.1

NE_DETECT  = re.compile(r'(^|\s)n/e(\s|$)', re.IGNORECASE)
NE_REMOVE  = re.compile(r'(^|\s)n/e(\s|$)', re.IGNORECASE)
LAYOFF     = re.compile(r'layoff', re.IGNORECASE)

# ── HEADER ALIASES ────────────────────────────────────────────────────────────

ALIASES = {
    'show':           ['show', 'showname', 'production'],
    'theatre':        ['theatre', 'theater', 'venue', 'theatrename'],
    'city':           ['city', 'market', 'location'],
    'ticketRange':    ['ticketrange', 'regticketrange', 'ticketpricerange'],
    'topPrice':       ['topprice', 'toppaidprice', 'top', 'premium'],
    'numPerf':        ['numperf', 'performances', 'perfs', 'perf'],
    'grossGross':     ['grossgross', 'gross', 'gg'],
    'grossPotential': ['grosspotential', 'potential', 'gp'],
    'ggPctGP':        ['ggpctgp', 'ggpgp', 'gpctgp', 'gggp'],
    'paidTix':        ['paidtix', 'paidattn', 'paidattendance', 'paid', 'paidtickets', 'paidticket'],
    'totalTix':       ['totaltix', 'totalattn', 'totalattendance', 'total', 'totaltickets', 'totalticket'],
    'capacity':       ['capacity', 'totalcapacity', 'cap'],
    'capPaid':        ['cappaid', 'capacitypaid'],
    'capTotal':       ['captotal', 'capacitytotal'],
    'onSub':          ['onsub', 'sub', 'subscription'],
    'avgAdm':         ['avgadm', 'avgpaidadmission', 'avgpaid', 'averagepaidadmission'],
}

# ── HELPERS ───────────────────────────────────────────────────────────────────

def normalize_header(value):
    return re.sub(r'[^a-z0-9]', '', str(value or '').lower().replace('\n', ' '))


def build_lookup():
    lookup = {}
    for field, aliases in ALIASES.items():
        for alias in aliases:
            key = normalize_header(alias)
            if key not in lookup:
                lookup[key] = field
    return lookup


def map_columns(header_row):
    lookup = build_lookup()
    col_map = {}
    for i, cell in enumerate(header_row):
        key = normalize_header(cell)
        if key in lookup and lookup[key] not in col_map:
            col_map[lookup[key]] = i
    return col_map


def find_header(rows):
    """Return index of the header row (first row containing 'show' and 'theatre')."""
    for i, row in enumerate(rows[:5]):
        cells = [str(c or '').strip().lower().replace('\n', ' ') for c in row]
        if 'show' in cells and ('theatre' in cells or 'theater' in cells):
            return i
    return None


def extract_date(sheet_name):
    """Extract ISO date string (YYYY-MM-DD) from sheet name."""
    m = re.search(r'(\d{1,2})[-_](\d{1,2})[-_](\d{2,4})', sheet_name)
    if not m:
        return None
    mm = m.group(1).zfill(2)
    dd = m.group(2).zfill(2)
    yy = m.group(3)
    if len(yy) == 2:
        yy = '20' + yy
    return f"{yy}-{mm}-{dd}"


def parse_number(v):
    if v is None or v == '':
        return None
    if isinstance(v, (int, float)):
        return None if (v != v) else float(v)  # NaN check
    s = str(v).strip()
    if not s or s == '-' or s.lower() == 'n/a':
        return None
    neg = s.startswith('(') and s.endswith(')')
    cleaned = re.sub(r'[$%(),]', '', s).strip()
    try:
        n = float(cleaned)
        return -n if neg else n
    except ValueError:
        return None


def parse_percent(v):
    n = parse_number(v)
    if n is None:
        return None
    return n * 100 if 0 < n < 5 else n


def parse_bool(v):
    return str(v or '').strip().lower() in ('x', 'yes', 'y', 'true', '1')


def normalize_key(s):
    return re.sub(r'[^a-z0-9 ]', '', s.lower()).strip()


# ── ROW PROCESSING ────────────────────────────────────────────────────────────

def process_row(row, col_map, week_of, tier):
    def get(field):
        idx = col_map.get(field)
        return row[idx] if idx is not None and idx < len(row) else None

    def text(field):
        return str(get(field) or '').strip()

    raw_show    = text('show')
    raw_theatre = text('theatre')
    raw_city    = text('city')

    if not raw_show or not raw_city:
        return None

    # nonEquity: n/e present in show, theatre, or city
    non_equity = (
        bool(NE_DETECT.search(raw_show)) or
        bool(NE_DETECT.search(raw_theatre)) or
        bool(NE_DETECT.search(raw_city))
    )

    # Clean show name — strip n/e marker
    show = NE_REMOVE.sub(' ', raw_show)
    show = re.sub(r'\s{2,}', ' ', show).strip()

    if not show or 'for engagements' in show.lower():
        return None

    num_perf    = parse_number(get('numPerf'))
    gross_gross = parse_number(get('grossGross'))

    # noEngagement: layoff in any field OR missing perf/gross data
    no_engagement = (
        bool(LAYOFF.search(raw_show)) or
        bool(LAYOFF.search(raw_theatre)) or
        bool(LAYOFF.search(raw_city)) or
        num_perf is None or
        gross_gross is None
    )

    capacity       = parse_number(get('capacity'))
    venue_sellable = round(capacity / num_perf, 2) if capacity and num_perf else None

    similar_bushnell = (
        venue_sellable is not None and
        LOWER_BOUND <= venue_sellable <= UPPER_BOUND
    )

    canonical_key = '|'.join([
        week_of,
        normalize_key(show),
        normalize_key(raw_theatre),
        normalize_key(raw_city),
        tier.lower()
    ])

    return {
        'week_of':          week_of,
        'tier':             tier,
        'show':             show,
        'theatre':          raw_theatre,
        'city':             raw_city,
        'ticket_range':     text('ticketRange') or None,
        'top_price':        parse_number(get('topPrice')),
        'num_perf':         num_perf,
        'gross_gross':      gross_gross,
        'gross_potential':  parse_number(get('grossPotential')),
        'gg_pct_gp':        parse_percent(get('ggPctGP')),
        'paid_tix':         parse_number(get('paidTix')),
        'total_tix':        parse_number(get('totalTix')),
        'capacity':         capacity,
        'cap_paid':         parse_percent(get('capPaid')),
        'cap_total':        parse_percent(get('capTotal')),
        'on_sub':           parse_bool(get('onSub')),
        'avg_adm':          parse_number(get('avgAdm')),
        'venue_sellable':   venue_sellable,
        'similar_bushnell': similar_bushnell,
        'non_equity':       non_equity,
        'no_engagement':    no_engagement,
        'canonical_key':    canonical_key,
    }


# ── FILE PROCESSING ───────────────────────────────────────────────────────────

def process_file(filepath, log):
    records = []
    fname = os.path.basename(filepath)

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        log.append(f"ERROR  | {fname} | Could not open: {e}")
        return records

    for sname in wb.sheetnames:
        upper = sname.upper()
        if 'PRIMARY' not in upper and 'SECONDARY' not in upper:
            continue

        tier = 'Primary' if 'PRIMARY' in upper else 'Secondary'
        week_of = extract_date(sname)

        if not week_of:
            log.append(f"WARN   | {fname} | {sname} | No date found in sheet name — skipped")
            continue

        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))

        h_idx = find_header(rows)
        if h_idx is None:
            log.append(f"WARN   | {fname} | {sname} | No header row found — skipped")
            continue

        col_map = map_columns(rows[h_idx])
        missing = [f for f in ALIASES if f not in col_map]
        if missing:
            log.append(f"WARN   | {fname} | {sname} | Missing columns: {missing}")

        sheet_records = 0
        for row in rows[h_idx + 1:]:
            rec = process_row(row, col_map, week_of, tier)
            if rec:
                records.append(rec)
                sheet_records += 1

        log.append(f"OK     | {fname} | {sname} | {sheet_records} records | week={week_of}")

    return records


# ── DEDUPLICATION ─────────────────────────────────────────────────────────────

def deduplicate(records, log):
    seen = {}
    dupes = 0
    for rec in records:
        key = rec['canonical_key']
        if key not in seen:
            seen[key] = rec
        else:
            dupes += 1
    log.append(f"INFO   | Deduplication: {len(records)} in, {dupes} dupes removed, {len(seen)} out")
    return list(seen.values())


# ── MAIN — FULL REBUILD ───────────────────────────────────────────────────────

def main_rebuild(input_dir, output_file):
    """Process all XLSX files in a directory and write a fresh data.json."""
    if not os.path.isdir(input_dir):
        print(f"Error: '{input_dir}' is not a directory")
        sys.exit(1)

    xlsx_files = sorted([
        os.path.join(input_dir, f)
        for f in os.listdir(input_dir)
        if f.lower().endswith('.xlsx') and not f.startswith('~')
    ])

    if not xlsx_files:
        print(f"No .xlsx files found in '{input_dir}'")
        sys.exit(1)

    log = []
    log.append(f"INFO   | MODE: full rebuild")
    log.append(f"INFO   | Processing {len(xlsx_files)} files from '{input_dir}'")

    all_records = []
    for fpath in xlsx_files:
        records = process_file(fpath, log)
        all_records.extend(records)

    all_records = deduplicate(all_records, log)
    all_records.sort(key=lambda r: (r['week_of'], r['show']))

    output = {
        'generated_at': datetime.now(timezone.utc).isoformat(),
        'record_count': len(all_records),
        'records': all_records,
    }

    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    print('\n'.join(log))
    print(f"\nDone. {len(all_records)} records written to '{output_file}'")


# ── MAIN — APPEND MODE ────────────────────────────────────────────────────────

def main_append(new_file, data_json):
    """
    Append a single new weekly XLSX file into an existing data.json.
    Deduplicates by canonical key — existing records are never overwritten.

    Usage:
        python process_touring.py --append <new_file.xlsx> <data.json>
    """
    if not os.path.isfile(new_file):
        print(f"Error: '{new_file}' not found")
        sys.exit(1)

    if not os.path.isfile(data_json):
        print(f"Error: '{data_json}' not found")
        sys.exit(1)

    log = []
    log.append(f"INFO   | MODE: append")
    log.append(f"INFO   | New file: {os.path.basename(new_file)}")
    log.append(f"INFO   | Target: {data_json}")

    # Load existing data
    with open(data_json, 'r', encoding='utf-8') as f:
        existing = json.load(f)

    existing_records = existing if isinstance(existing, list) else existing.get('records', [])
    existing_keys = {r['canonical_key'] for r in existing_records}
    log.append(f"INFO   | Existing records: {len(existing_records)}")

    # Process new file
    new_records = process_file(new_file, log)

    # Append only records not already present
    added = 0
    dupes = 0
    for rec in new_records:
        if rec['canonical_key'] not in existing_keys:
            existing_records.append(rec)
            existing_keys.add(rec['canonical_key'])
            added += 1
        else:
            dupes += 1

    log.append(f"INFO   | New records added: {added} | Duplicates skipped: {dupes}")

    # Re-sort
    existing_records.sort(key=lambda r: (r['week_of'], r['show']))

    output = {
        'generated_at': datetime.now(timezone.utc).isoformat(),
        'record_count': len(existing_records),
        'records': existing_records,
    }

    with open(data_json, 'w', encoding='utf-8') as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    print('\n'.join(log))
    print(f"\nDone. {added} records added. Total: {len(existing_records)} records in '{data_json}'")


# ── ENTRY POINT ───────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) >= 2 and sys.argv[1] == '--append':
        if len(sys.argv) != 4:
            print("Usage: python process_touring.py --append <new_file.xlsx> <data.json>")
            sys.exit(1)
        main_append(sys.argv[2], sys.argv[3])
    else:
        if len(sys.argv) != 3:
            print("Usage: python process_touring.py <input_dir> <output_file>")
            print("       python process_touring.py --append <new_file.xlsx> <data.json>")
            sys.exit(1)
        main_rebuild(sys.argv[1], sys.argv[2])


if __name__ == '__main__':
    main()
